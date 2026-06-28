import os
from datetime import datetime
from flask import Blueprint, request, send_file, current_app
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models.patient import Patient
from app.models.appointment import Appointment
from app.models.medical_record import MedicalRecord
from app.schemas.medical_record import ExportQuerySchema
from app.utils.errors import ApiException, success_response
from app.utils.business import parse_date

bp = Blueprint('exports', __name__)


def ensure_export_dir():
    base_dir = os.path.abspath(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    export_dir = os.path.join(base_dir, 'exports')
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    return export_dir


def set_cell_style(cell, is_header=False, is_weekend=False):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if is_header:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, size=11, color='FFFFFF')
    elif is_weekend:
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')


@bp.route('/exports/patients', methods=['GET'])
def export_patients():
    export_dir = ensure_export_dir()

    patients = Patient.query.order_by(Patient.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '患者列表'

    headers = ['编号', '姓名', '手机号', '性别', '年龄', '住址', '病史备注', '建档时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_cell_style(cell, is_header=True)

    widths = [8, 12, 15, 8, 8, 30, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    status_map = {'pending': '待确认', 'confirmed': '已确认', 'completed': '已完成', 'cancelled': '已取消'}

    for row_idx, patient in enumerate(patients, 2):
        values = [
            patient.id,
            patient.name,
            patient.phone,
            patient.gender or '',
            patient.age or '',
            patient.address or '',
            patient.medical_history or '',
            patient.created_at.strftime('%Y-%m-%d %H:%M:%S') if patient.created_at else ''
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            set_cell_style(cell)

    filename = f'患者列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/exports/appointments', methods=['GET'])
def export_appointments():
    schema = ExportQuerySchema()
    try:
        params = schema.load(request.args.to_dict())
    except Exception as e:
        raise ApiException(str(e), 400)

    export_dir = ensure_export_dir()

    query = Appointment.query
    if params.get('start_date'):
        sd = parse_date(params['start_date'])
        if sd:
            query = query.filter(Appointment.appointment_date >= sd)
    if params.get('end_date'):
        ed = parse_date(params['end_date'])
        if ed:
            query = query.filter(Appointment.appointment_date <= ed)
    if params.get('patient_id'):
        query = query.filter_by(patient_id=params['patient_id'])
    if params.get('doctor_id'):
        query = query.filter_by(doctor_id=params['doctor_id'])

    appointments = query.order_by(
        Appointment.appointment_date,
        Appointment.doctor_id,
        Appointment.appointment_time
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '预约记录'

    headers = ['编号', '预约日期', '预约时段', '星期', '医生姓名', '患者姓名', '手机号', '状态', '主诉', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_cell_style(cell, is_header=True)

    widths = [8, 14, 10, 8, 12, 12, 15, 10, 25, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    status_map = {'pending': '待确认', 'confirmed': '已确认', 'completed': '已完成', 'cancelled': '已取消'}

    for row_idx, appt in enumerate(appointments, 2):
        is_weekend = appt.appointment_date.weekday() >= 5
        values = [
            appt.id,
            appt.appointment_date.strftime('%Y-%m-%d'),
            appt.appointment_time,
            weekdays[appt.appointment_date.weekday()],
            appt.doctor.name if appt.doctor else '',
            appt.patient.name if appt.patient else '',
            appt.patient.phone if appt.patient else '',
            status_map.get(appt.status, appt.status),
            appt.chief_complaint or '',
            appt.notes or ''
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            set_cell_style(cell, is_weekend=is_weekend)

    date_suffix = ''
    if params.get('start_date'):
        date_suffix += f'_{params["start_date"]}'
    if params.get('end_date'):
        date_suffix += f'_{params["end_date"]}'
    filename = f'预约记录{date_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/exports/records', methods=['GET'])
def export_records():
    schema = ExportQuerySchema()
    try:
        params = schema.load(request.args.to_dict())
    except Exception as e:
        raise ApiException(str(e), 400)

    export_dir = ensure_export_dir()

    query = MedicalRecord.query
    if params.get('start_date'):
        sd = parse_date(params['start_date'])
        if sd:
            query = query.filter(MedicalRecord.visit_date >= sd)
    if params.get('end_date'):
        ed = parse_date(params['end_date'])
        if ed:
            query = query.filter(MedicalRecord.visit_date <= ed)
    if params.get('patient_id'):
        query = query.filter_by(patient_id=params['patient_id'])
    if params.get('doctor_id'):
        query = query.filter_by(doctor_id=params['doctor_id'])

    records = query.order_by(
        MedicalRecord.visit_date,
        MedicalRecord.doctor_id,
        MedicalRecord.created_at
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '诊疗记录'

    headers = ['编号', '就诊日期', '医生姓名', '患者姓名', '手机号', '诊断', '治疗方案', '处方', '费用(元)', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_cell_style(cell, is_header=True)

    widths = [8, 14, 12, 12, 15, 35, 35, 30, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    total_fees = 0
    for row_idx, record in enumerate(records, 2):
        fees = float(record.fees or 0)
        total_fees += fees
        values = [
            record.id,
            record.visit_date.strftime('%Y-%m-%d'),
            record.doctor.name if record.doctor else '',
            record.patient.name if record.patient else '',
            record.patient.phone if record.patient else '',
            record.diagnosis or '',
            record.treatment or '',
            record.prescription or '',
            fees,
            record.notes or ''
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            set_cell_style(cell)

    summary_row = len(records) + 3
    ws.cell(row=summary_row, column=1, value='合计').font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value='总费用(元)').font = Font(bold=True)
    ws.cell(row=summary_row, column=9, value=total_fees).font = Font(bold=True, color='C00000')

    date_suffix = ''
    if params.get('start_date'):
        date_suffix += f'_{params["start_date"]}'
    if params.get('end_date'):
        date_suffix += f'_{params["end_date"]}'
    filename = f'诊疗记录{date_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/exports/summary', methods=['GET'])
def export_daily_summary():
    date_str = request.args.get('date')
    if not date_str:
        raise ApiException('请指定查询日期', 400)

    target_date = parse_date(date_str)
    if not target_date:
        raise ApiException('日期格式不正确', 400)

    export_dir = ensure_export_dir()
    weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    status_map = {'pending': '待确认', 'confirmed': '已确认', 'completed': '已完成', 'cancelled': '已取消'}

    wb = Workbook()
    ws = wb.active
    ws.title = '营业日报'

    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value=f'{date_str} 营业日报（{weekdays[target_date.weekday()]}）')
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    from app.models.doctor import Doctor
    doctors = Doctor.query.filter_by(is_active=True).all()

    headers = ['时段', '星期', '医生', '状态', '患者姓名', '手机号', '主诉', '费用']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        set_cell_style(cell, is_header=True)

    widths = [14, 8, 12, 10, 12, 15, 25, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    row_idx = 4
    total_appointments = 0
    total_fees = 0
    is_weekend = target_date.weekday() >= 5

    for doctor in doctors:
        appts = Appointment.query.filter(
            Appointment.doctor_id == doctor.id,
            Appointment.appointment_date == target_date
        ).order_by(Appointment.appointment_time).all()

        total_appointments += len([a for a in appts if a.status != 'cancelled'])

        for appt in appts:
            fees = 0
            if appt.medical_records:
                fees = sum(float(r.fees or 0) for r in appt.medical_records)
            total_fees += fees

            values = [
                f'{appt.appointment_date.strftime("%Y-%m-%d")} {appt.appointment_time}',
                weekdays[appt.appointment_date.weekday()],
                doctor.name,
                status_map.get(appt.status, appt.status),
                appt.patient.name if appt.patient else '',
                appt.patient.phone if appt.patient else '',
                appt.chief_complaint or '',
                fees
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                set_cell_style(cell, is_weekend=is_weekend)
            row_idx += 1

    summary_row = row_idx + 2
    ws.cell(row=summary_row, column=1, value='当日预约数').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=total_appointments).font = Font(bold=True, color='C00000')
    ws.cell(row=summary_row, column=6, value='当日诊疗费(元)').font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value=total_fees).font = Font(bold=True, color='C00000')

    filename = f'营业日报_{date_str}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
